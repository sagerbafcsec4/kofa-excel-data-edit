#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
サッカー記録ワークブック バッチ整形スクリプト (excel-data skill 準拠)

★このファイルがロジックの唯一の正本。Webアプリ(index.html)は実行時にこれを読み込むため、
  ロジックの修正はこのファイルだけ直せばよい(index.html 側に書き写す必要はない)。
"""
# redeploy: 2026-07-04 13:50
import sys, os, glob
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
from collections import Counter

# --- 全角カタカナ→半角カタカナ 変換(表記まとめのタブ名用) ---
_HALF = {}
_bf = "アイウエオカキクケコサシスセソタチツテトナニヌネノハヒフヘホマミムメモヤユヨラリルレロワヲン"
_bh = "ｱｲｳｴｵｶｷｸｹｺｻｼｽｾｿﾀﾁﾂﾃﾄﾅﾆﾇﾈﾉﾊﾋﾌﾍﾎﾏﾐﾑﾒﾓﾔﾕﾖﾗﾘﾙﾚﾛﾜｦﾝ"
for _f, _h in zip(_bf, _bh):
    _HALF[_f] = _h
for _f, _h in zip("ァィゥェォッャュョ", "ｧｨｩｪｫｯｬｭｮ"):
    _HALF[_f] = _h
for _f, _b in zip("ガギグゲゴザジズゼゾダヂヅデドバビブベボ", "カキクケコサシスセソタチツテトハヒフヘホ"):
    _HALF[_f] = _HALF[_b] + "ﾞ"
for _f, _b in zip("パピプペポ", "ハヒフヘホ"):
    _HALF[_f] = _HALF[_b] + "ﾟ"
_HALF["ヴ"] = _HALF["ウ"] + "ﾞ"
_HALF["ー"] = "ｰ"; _HALF["・"] = "･"; _HALF["　"] = " "

def to_half_kana(s):
    return "".join(_HALF.get(ch, ch) for ch in str(s))

try:
    BASE = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE = os.getcwd()
OUT  = os.path.join(BASE, "_processed")
LOGS = os.path.join(BASE, "_logs")
DELETE_KEIREKI = None
MATCH_DATE = None       # 試合日 "YYYY/MM/DD"。None(未指定)なら歳列は変更しない
DELETE_WEIGHT = None    # True なら表記シートの体重(kg)列を削除
AGE_NOTE = None         # 「歳」列下の注記文言。None ならデフォルト文言を使用

def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def jwidth(text):
    w = 0
    for ch in str(text):
        w += 2 if ord(ch) > 0x2E7F else 1
    return w

import unicodedata
def dispwidth(text):
    """実表示幅(半角=1/全角=2)。半角カタカナ等を正しく1と数える(jwidthは2に数える)。"""
    w = 0
    for ch in str(text):
        w += 2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
    return w

def set_font_size(cell, size):
    f = copy(cell.font); f.size = size; cell.font = f

def set_shrink(cell):
    a = copy(cell.alignment); a.shrink_to_fit = True; cell.alignment = a

def read_grid(ws):
    return [list(row) for row in ws.iter_rows(values_only=True)]

def g(grid, r, c):
    if 1 <= r <= len(grid) and 1 <= c <= len(grid[r-1]):
        return grid[r-1][c-1]
    return None

def last_data_row(grid, c1, c2):
    last = 1
    for r in range(1, len(grid) + 1):
        row = grid[r-1]
        for c in range(c1, min(c2, len(row)) + 1):
            if not is_empty(row[c-1]):
                last = r; break
    return last

def col_empty(grid, c, r1, r2):
    for r in range(r1, r2 + 1):
        if not is_empty(g(grid, r, c)):
            return False
    return True

WIDTHS = {"節":10,"開催日":10,"H":3,"A":3,"ｽｺｱ":8,"スコア":8,"対戦相手":30,"退場":20,"NEWS":60}

def _is_note(v):
    """表外の注記セルか(縮小・書式変更してはいけない)。"""
    return isinstance(v, str) and ("用語説明" in v or v.lstrip().startswith("【"))

def fmt_season(ws, log):
    grid = read_grid(ws)
    last = last_data_row(grid, 1, 18)
    for cd in ws.column_dimensions.values():
        cd.hidden = False
    score_cols = [c for c in range(1, 19)
                  if isinstance(g(grid, 2, c), str) and g(grid, 2, c).strip().startswith("得点")]
    del_score = []
    if score_cols:
        start = score_cols[0]; block = list(range(start, start + 8))
        del_score = [block[i] for i in range(len(block) - 1, 2, -1) if col_empty(grid, block[i], 3, last)]
    del_empty = [c for c in range(1, 19) if is_empty(g(grid, 2, c)) and col_empty(grid, c, 2, last)]
    for c in sorted(set(del_score) | set(del_empty), reverse=True):
        lab = "得点空" if c in del_score else "空ヘッダー"
        log(f"  {get_column_letter(c)}({lab}) -> 削除")
        ws.delete_cols(c, 1)
    grid = read_grid(ws)
    fmax = max((c for c in range(1, 26) if not is_empty(g(grid, 2, c))), default=18)
    # フォント11pt + 縮小は「表の枠内」だけ。表外の注記(【用語説明】等)は触らない。
    tbl_last = max((r for r in range(3, last + 1) if not is_empty(g(grid, r, 1))), default=2)
    for row in ws.iter_rows(min_row=2, max_row=tbl_last, min_col=1, max_col=fmax):
        for cell in row:
            if _is_note(cell.value):
                continue
            set_font_size(cell, 11); set_shrink(cell)
    # 表外の注記(用語説明等)に残っている「縮小して全体を表示」を能動的に解除する
    for row in ws.iter_rows():
        for cell in row:
            if _is_note(cell.value) and cell.alignment.shrink_to_fit:
                a = copy(cell.alignment); a.shrink_to_fit = False; cell.alignment = a
                log(f"  注記 {cell.coordinate} の縮小を解除")
    for c in range(1, fmax + 1):
        h = g(grid, 2, c)
        h = h.strip() if isinstance(h, str) else h
        if h in WIDTHS:
            w = WIDTHS[h]
        elif isinstance(h, str) and (h.startswith("得点") or h.isdigit()):
            w = 20
        else:
            continue
        ws.column_dimensions[get_column_letter(c)].width = w
        log(f"  幅 {get_column_letter(c)}('{h}') -> {w}")

PLAY_POS = {"GK","DF","MF","FW"}

def fmt_appearance(ws, log):
    grid = read_grid(ws)
    H = 8
    maxc = max((len(r) for r in grid), default=0)
    last = last_data_row(grid, 1, maxc)
    # 左上のチーム名セル(結合 A1:C2 等)は列幅を広げず、必ず「縮小して全体を表示する」を有効化。
    # (長いチーム名がはみ出て切れるのを防ぐ。値は変えない=書式のみ)
    title = ws.cell(1, 1)
    if not is_empty(title.value):
        a = copy(title.alignment)
        a.shrink_to_fit = True
        a.wrap_text = False   # 縮小表示と折り返しは排他。折り返しは無効化する。
        title.alignment = a
        log(f"  チーム名 A1 に『縮小して全体を表示』を設定")
    # C列(名前): Excelのオートフィット(列境界ダブルクリック)相当の幅にする。
    # 狭めるの禁止=現在幅より必要幅が大きいときだけ広げる。
    # オートフィット幅 ≈ セルごとの「文字幅(全角=2/半角=1) × フォントpt ÷ 11」の最大値。
    #   (Excelの列幅単位はMDW≒7pxに基づき、この式で実測(例: 全角42×16pt→61)と一致する)
    need = 0.0
    for r in range(1, last + 1):
        cell = ws.cell(r, 3)
        v = cell.value
        if is_empty(v):
            continue
        sz = cell.font.size or 11
        need = max(need, jwidth(v) * sz / 11.0)
    if need:
        truew = {}
        for k, cd in list(ws.column_dimensions.items()):
            if cd.width is None: continue
            mn = cd.min or column_index_from_string(k); mx = cd.max or mn
            for x in range(mn, mx + 1): truew[x] = cd.width
        curC = truew.get(3, 8.43)
        target = min(round(need), 100)
        if target > curC + 1:
            ws.column_dimensions["C"].width = target
            log(f"  名前列C {round(curC,2)}->{target} (オートフィット相当に拡張)")
    def match_empty_2row(r1, r2):
        for r in (r1, r2):
            for c in range(H, maxc + 1):
                if not is_empty(g(grid, r, c)):
                    return False
        return True
    del_players = []
    r = 6
    while r <= last:
        pos = g(grid, r, 1)
        pos_s = pos.strip() if isinstance(pos, str) else pos
        if pos_s in PLAY_POS:
            if is_empty(g(grid, r, 2)) and match_empty_2row(r, r+1):
                del_players.append((r, g(grid, r, 3)))
            r += 2
        else:
            r += 1
    for r, name in sorted(del_players, reverse=True):
        log(f"  記録ゼロ選手 行{r}-{r+1} '{name}' -> 2行削除")
        ws.delete_rows(r, 2)
    grid2 = read_grid(ws)
    maxc2 = max((len(r) for r in grid2), default=0)
    last2 = last_data_row(grid2, 1, maxc2)
    def match_unplayed(c):
        if not is_empty(g(grid2, 4, c)):
            return False
        for cc in (c, c+1):
            for r in range(6, last2 + 1):
                if not is_empty(g(grid2, r, cc)):
                    return False
        return True
    leftmost = None
    c = H
    while c <= maxc2:
        if match_unplayed(c):
            leftmost = c; break
        c += 2
    if leftmost is not None:
        cut = leftmost + 2
        if cut <= maxc2:
            n = maxc2 - cut + 1
            log(f"  未実施試合 {get_column_letter(cut)}:{get_column_letter(maxc2)} ({n}列) -> 削除")
            ws.delete_cols(cut, n)
    # 凡例(◎：フル出場…)の結合は指示外なので解除する(最終レイアウトで判定)
    g3 = read_grid(ws)
    legend_row = next((r for r in range(1, len(g3) + 1)
                       if isinstance(g(g3, r, 1), str) and "フル出場" in g(g3, r, 1)), None)
    if legend_row is not None:
        removed = 0
        for m in list(ws.merged_cells.ranges):
            if m.min_row == legend_row:
                ws.unmerge_cells(str(m)); removed += 1
        if removed:
            log(f"  凡例(行{legend_row})の結合を {removed}件 解除")
    # 監督・AC(A列ラベル)のA・B結合を保持/付与: 行削除でA-B横結合が割れることがあるので再結合する。
    # ACも監督と同様にA列・B列を結合する。
    for label in ("監督", "AC"):
        lr = next((r for r in range(1, len(g3) + 1)
                   if str(g(g3, r, 1)).strip() == label), None)
        if lr is None:
            continue
        bottom = lr + 1
        for m in list(ws.merged_cells.ranges):
            if m.min_col <= 2 and m.min_row <= lr <= m.max_row:
                bottom = max(bottom, m.max_row); ws.unmerge_cells(str(m))
        ws.merge_cells(start_row=lr, start_column=1, end_row=bottom, end_column=2)
        log(f"  {label}のA・B結合を保持 (A{lr}:B{bottom})")
    # 印刷範囲: 一番下を凡例行に。右端は内容のある最終列。
    if legend_row is not None:
        last_col = 1
        for r in range(1, legend_row + 1):
            for c in range(1, (len(g3[r-1]) if r <= len(g3) else 0) + 1):
                if not is_empty(g(g3, r, c)):
                    last_col = max(last_col, c)
        ws.print_area = f"A1:{get_column_letter(last_col)}{legend_row}"
        log(f"  印刷範囲 A1:{get_column_letter(last_col)}{legend_row}")

PANELS = [(2,25),(26,49),(50,73),(74,97)]
RIGHT_COL = {1:"Y", 2:"AW", 3:"BU", 4:"CS"}
FILLED_MIN = 40

def fmt_formation(ws, wb, log):
    grid = read_grid(ws)
    n = 0
    for r1, r2 in ((1,49),(50,97)):
        for c1, c2 in PANELS:
            cnt = 0
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    if not is_empty(g(grid, r, c)):
                        cnt += 1
                if cnt >= FILLED_MIN: break
            if cnt >= FILLED_MIN: n += 1
    log(f"  フォーメーション数 = {n}")
    if n == 0:
        log(f"  -> 空シート '{ws.title}' 削除"); wb.remove(ws); return
    if n <= 4:
        # 上段の下端を動的に決定。テンプレの「負傷」行が上段/下段に1つずつあるので、
        # その間隔=バンド高さ。上段の下端=1+バンド高さ → スタッツ(枠/CK)行まで含まれる。
        injuries = [r for r in range(1, len(grid) + 1)
                    if any(isinstance(g(grid, r, c), str) and g(grid, r, c).strip() == "負傷"
                           for c in (2, 26, 50, 74))]
        bottom = (1 + (injuries[1] - injuries[0])) if len(injuries) >= 2 else 49
        right = RIGHT_COL[n]
        ws.print_area = f"A1:{right}{bottom}"
        log(f"  -> 印刷範囲 A1:{right}{bottom}")
    else:
        log(f"  -> 5試合以上、印刷範囲は変更なし")

def fmt_hyoki(ws, log):
    grid = read_grid(ws)
    maxc = max((len(r) for r in grid), default=0)
    # 体重(kg)列の削除(オプション)。行2ヘッダーが「kg」の列を1つ削除する。
    if DELETE_WEIGHT:
        kg_col = None
        for c in range(1, maxc + 1):
            h = g(grid, 2, c)
            if isinstance(h, str) and h.strip() == "kg":
                kg_col = c; break
        if kg_col:
            log(f"  体重列 {get_column_letter(kg_col)}('kg') -> 削除")
            # 削除前に全列の実幅(範囲展開)を保存
            pre_w = {}
            for k, cd in list(ws.column_dimensions.items()):
                if cd.width is None:
                    continue
                mn = cd.min or column_index_from_string(k)
                mx = cd.max or mn
                for x in range(mn, mx + 1):
                    pre_w[x] = cd.width
            old_maxc = maxc
            ws.delete_cols(kg_col, 1)
            # 重要: openpyxl は列削除時に<col>(列幅)定義をずらさないため、
            #   削除後の各列にデータ追従した幅を手動で再割当する(そうしないと
            #   試/国籍などが右隣の幅を引き継いで広がる/狭まるバグになる)。
            for k in list(ws.column_dimensions.keys()):
                del ws.column_dimensions[k]
            for c in range(1, old_maxc):      # 削除後の有効列数 = old_maxc - 1
                src = c if c < kg_col else c + 1
                if src in pre_w:
                    ws.column_dimensions[get_column_letter(c)].width = pre_w[src]
            grid = read_grid(ws)
            maxc = max((len(r) for r in grid), default=0)
        else:
            log("  体重列(kg) が見つからないため削除なし")
    last = last_data_row(grid, 1, maxc)
    age_col = bd_col = None
    for c in range(1, maxc + 1):
        h = g(grid, 2, c)
        h = h.strip() if isinstance(h, str) else h
        if h == "歳": age_col = c
        elif h == "生年月日": bd_col = c
    if MATCH_DATE and age_col and bd_col:
        bdL = get_column_letter(bd_col)
        date_str = MATCH_DATE.replace("-", "/")
        cnt_age = 0
        for r in range(3, last + 1):
            if not is_empty(g(grid, r, bd_col)):
                ws.cell(r, age_col).value = f'=DATEDIF({bdL}{r}, "{date_str}", "Y")'
                cnt_age += 1
        log(f"  歳列({get_column_letter(age_col)}) に {date_str} 基準の年齢式を {cnt_age}件 設定")
        note_text = str(AGE_NOTE if AGE_NOTE is not None else "※年齢は試合当日のもの").strip()
        if note_text:
            ws.cell(last + 1, age_col).value = note_text
            log(f"  注記 {get_column_letter(age_col)}{last + 1} に『{note_text}』を設定")
    elif age_col:
        log("  歳列: 試合日未指定のため変更なし")
    # 列幅(表記のみ): 狭めるの禁止。広げるのは文字が切れる(内容>現在幅)時だけ。タイトル(行1)は除外。
    # 重要: openpyxlは範囲指定された<col>(例 A-B共通幅)を持つ列の column_dimensions に
    #   アクセス/設定すると勝手に幅を書き換えてしまう。よって
    #   (1) 既存の範囲を展開して各列の"本当の現在幅"を求める
    #   (2) 実際に広げる列だけ column_dimensions を触る(それ以外は一切触らない=原本保持)
    DEFAULT_W = 8.43
    truew = {}
    for key, cd in list(ws.column_dimensions.items()):
        if cd.width is None:
            continue
        mn = cd.min or column_index_from_string(key)
        mx = cd.max or mn
        for ci in range(mn, mx + 1):
            truew[ci] = cd.width
    # Excelのオートフィット(列境界ダブルクリック)相当で「広げるか」を判断する。
    # 幅 ≈ セルごとの「文字幅(全角=2/半角=1) × フォントpt ÷ 11」の最大値。
    # 値は元データ(grid)から取得する=歳列に入れた数式や注記の文字数に引きずられない。
    for c in range(1, maxc + 1):
        need = 0.0
        for r in range(2, last + 1):
            v = g(grid, r, c)
            if is_empty(v):
                continue
            sz = ws.cell(r, c).font.size or 11
            need = max(need, jwidth(v) * sz / 11.0)
        if need == 0:
            continue
        cur = truew.get(c, DEFAULT_W)
        target = min(round(need), 100)
        if target > cur + 1:   # オートフィット必要幅が現在より大きい列だけ広げる。狭めない。
            ws.column_dimensions[get_column_letter(c)].width = target
            log(f"  幅 {get_column_letter(c)} {round(cur,2)}->{target} (オートフィット相当に拡張)")

def fmt_keireki(ws, log):
    """経歴シートの整形(削除しない場合)。値は変更せず、書式/幅/高さ/印刷のみ。
    - 印刷 A4->A3(用紙のみ)
    - A1: フォント15・太字
    - 枠内(2行目〜監督行 × A〜G列, 空セル含む): フォント12 + 縮小&折り返し 両方ON
    - 行の高さ 1行目〜監督行 = 30
    - 列幅 A5 B5 C50 D5 E6 F50 G200
    監督行が無ければ最終データ行(A〜G)を下端にする。
    """
    grid = read_grid(ws)
    # 監督行(A列=='監督')、無ければ最終データ行
    mr = next((r for r in range(1, len(grid) + 1)
               if str(g(grid, r, 1)).strip() == "監督"), None)
    if mr is None:
        mr = last_data_row(grid, 1, 7)
    if mr < 2:
        mr = 2
    # 印刷: A4(9)->A3(8) 用紙サイズのみ
    ws.page_setup.paperSize = 8
    log("  用紙 A4->A3")
    # A1: 15pt 太字
    a1 = ws.cell(1, 1)
    f = copy(a1.font); f.size = 15; f.bold = True; a1.font = f
    log("  A1 フォント15・太字")
    # 枠内(2..mr × A..G): フォント12 + 縮小&折り返し
    for r in range(2, mr + 1):
        for c in range(1, 8):
            cell = ws.cell(r, c)
            ff = copy(cell.font); ff.size = 12; cell.font = ff
            aa = copy(cell.alignment)
            aa.shrink_to_fit = True
            aa.wrap_text = True
            cell.alignment = aa
    log(f"  枠内(2〜{mr}行 × A〜G列) フォント12＋縮小＋折り返し")
    # 行の高さ 1..mr = 30
    for r in range(1, mr + 1):
        ws.row_dimensions[r].height = 30
    log(f"  行の高さ 1〜{mr} = 30")
    # 経歴(G列)が長く折り返して3行以上になる行は、行の高さを 30->50 にする。
    # 1行あたり容量(表示幅) = G幅(200) * 11 / フォント(12) ≒ 183。3行以上=表示幅>約367。
    import math
    kei_col = None
    for c in range(1, 8):
        h = g(grid, 2, c)
        if isinstance(h, str) and h.strip() == "経歴":
            kei_col = c
    if kei_col:
        cap = 200 * 11 / 12.0
        n50 = 0
        for r in range(3, mr + 1):
            v = g(grid, r, kei_col)
            if is_empty(v):
                continue
            lines = math.ceil(dispwidth(v) / cap)
            if lines >= 3:
                ws.row_dimensions[r].height = 50
                n50 += 1
        log(f"  経歴が3行以上の行を 行高50 に: {n50}件")
    # 列幅
    for col, w in {"A": 5, "B": 5, "C": 50, "D": 5, "E": 6, "F": 50, "G": 200}.items():
        ws.column_dimensions[col].width = w
    log("  列幅 A5 B5 C50 D5 E6 F50 G200")
    # 歳(D)列に「その日の年齢」式を入れる(試合日指定時のみ)。
    # 経歴には生年月日が無いため、表記シートの生年月日を「フルネーム」で参照して計算する。
    # ※値を変更してよいのはこの歳列だけ(安全チェックでも歳列のみ照合除外)。
    if MATCH_DATE:
        kf = kage = None
        for c in range(1, 8):
            h = g(grid, 2, c)
            h = h.strip() if isinstance(h, str) else h
            if h == "フルネーム": kf = c
            elif h == "歳": kage = c
        hy = next((w for w in ws.parent.worksheets if is_hyoki_sheet(w)), None)
        hf = hbd = None
        if hy is not None:
            for c in range(1, (hy.max_column or 0) + 1):
                v = hy.cell(2, c).value
                v = v.strip() if isinstance(v, str) else v
                if v == "フルネーム": hf = c
                elif v == "生年月日": hbd = c
        if kf and kage and hy is not None and hf and hbd:
            date_str = MATCH_DATE.replace("-", "/")
            sref = "'" + str(hy.title).replace("'", "''") + "'"
            kfL = get_column_letter(kf); hfL = get_column_letter(hf); hbdL = get_column_letter(hbd)
            cnt = 0
            for r in range(3, mr + 1):
                if is_empty(g(grid, r, kage)) or is_empty(g(grid, r, kf)):
                    continue
                ws.cell(r, kage).value = (
                    f'=IFERROR(DATEDIF(INDEX({sref}!{hbdL}:{hbdL},'
                    f'MATCH({kfL}{r},{sref}!{hfL}:{hfL},0)),"{date_str}","Y"),"")'
                )
                cnt += 1
            log(f"  歳列({get_column_letter(kage)}) に {date_str} 基準の年齢式を {cnt}件 設定(表記の生年月日を氏名参照)")
        else:
            log("  歳列: 表記/フルネーム/生年月日 列が特定できず年齢式は未設定")

def _row_set(ws, r, cmax=14):
    s = set()
    for c in range(1, cmax + 1):
        v = ws.cell(r, c).value
        if v not in (None, "") and str(v).strip() != "":
            s.add(str(v).strip())
    return s

def _cell(ws, r, c):
    v = ws.cell(r, c).value
    return str(v).strip() if v not in (None, "") else ""

def is_appearance_sheet(ws):
    if _cell(ws, 1, 4) == "出場記録":
        return True
    return _cell(ws, 4, 1) == "Pos" and _cell(ws, 4, 2) == "No" and _cell(ws, 4, 3) == "Name"

def is_season_sheet(ws):
    h = _row_set(ws, 2)
    return "節" in h and "開催日" in h and "対戦相手" in h

def is_hyoki_sheet(ws):
    return "統一表記" in _row_set(ws, 2)

def is_keireki_sheet(ws):
    h = _row_set(ws, 2)
    return "主な下部組織" in h and "経歴" in h

def is_formation_sheet(ws):
    return 90 <= ws.max_column <= 110

def _clean_tab(name, fallback="表記"):
    """Excelタブ名として使える形に整える(禁則文字除去・前後空白除去・31字まで)。"""
    name = str(name)
    for bad in ":\\/?*[]":
        name = name.replace(bad, "")
    name = name.strip()[:31]
    return name or fallback

def hyoki_tab_name(ws):
    """表記シートA1のチーム名(「(」の前)を取り出し、カタカナは半角化してタブ名にする。"""
    a1 = ws.cell(1, 1).value
    name = str(a1) if a1 not in (None, "") else ""
    idx = len(name)
    for br in ("（", "("):
        p = name.find(br)
        if p != -1:
            idx = min(idx, p)
    return _clean_tab(to_half_kana(name[:idx].strip()))

def copy_ws_into(src, dst_wb, title):
    """src ワークシートを dst_wb に title で複製(値+主要書式+列幅+結合を保持)。"""
    dst = dst_wb.create_sheet(title=title)
    # シート既定書式(既定列幅/既定行高)を引き継ぐ。
    # (既定列幅を引き継がないと、明示幅の無い列が標準8.43に変わってしまう)
    try:
        dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
        dst.sheet_format.baseColWidth = src.sheet_format.baseColWidth
        dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    except Exception:
        pass
    sw = {}
    for k, cd in src.column_dimensions.items():
        if cd.width is None:
            continue
        mn = cd.min or column_index_from_string(k)
        mx = cd.max or mn
        for x in range(mn, mx + 1):
            sw[x] = cd.width
    for c, w in sw.items():
        dst.column_dimensions[get_column_letter(c)].width = w
    for idx, rd in src.row_dimensions.items():
        if rd.height is not None:
            dst.row_dimensions[idx].height = rd.height
    for row in src.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.fill = copy(cell.fill)
                nc.border = copy(cell.border)
                nc.alignment = copy(cell.alignment)
                nc.number_format = cell.number_format
                nc.protection = copy(cell.protection)
    for mc in list(src.merged_cells.ranges):
        dst.merge_cells(str(mc))
    if src.print_area:
        dst.print_area = src.print_area
    return dst

def add_hyoki_to_summary(wb, dst_wb, tab_name=None):
    """処理済み wb の表記シートを、まとめブック dst_wb に1タブ追加する。タブ名重複は連番。
    tab_name(ファイル名コード等)が渡ればそれを、無ければA1のチーム名を使う。"""
    ws = next((w for w in wb.worksheets if is_hyoki_sheet(w)), None)
    if ws is None:
        return None
    base = _clean_tab(tab_name) if (tab_name and str(tab_name).strip()) else hyoki_tab_name(ws)
    title = base
    i = 2
    while title in dst_wb.sheetnames:
        suffix = "_" + str(i)
        title = base[:31 - len(suffix)] + suffix
        i += 1
    copy_ws_into(ws, dst_wb, title)
    return title

def process_wb(wb, log):
    for ws in list(wb.worksheets):
        t = ws.title
        if is_appearance_sheet(ws):
            log(f"[§A 出場記録] {t}"); fmt_appearance(ws, log)
        elif is_season_sheet(ws):
            log(f"[§B シーズン] {t}"); fmt_season(ws, log)
        elif is_hyoki_sheet(ws):
            log(f"[§D 表記] {t}"); fmt_hyoki(ws, log)
        elif is_keireki_sheet(ws):
            if DELETE_KEIREKI:
                log(f"[§E 経歴] {t} -> 削除"); wb.remove(ws)
            else:
                log(f"[§E 経歴] {t} -> 整形保持"); fmt_keireki(ws, log)
        elif is_formation_sheet(ws):
            log(f"[§C フォーメーション] {t}"); fmt_formation(ws, wb, log)
        else:
            log(f"[--] {t} -> 対象外シート、スキップ")
    return wb

def _exempt_cols(ws):
    # 試合日指定時、値変更を許可する列=表記/経歴シートの「歳」列のみ。
    if MATCH_DATE and (is_hyoki_sheet(ws) or is_keireki_sheet(ws)):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(2, c).value
            if isinstance(v, str) and v.strip() == "歳":
                return {c}
    return set()

def _sheet_values(ws):
    exempt = _exempt_cols(ws)
    c = Counter()
    for row in ws.iter_rows(values_only=True):
        for idx, v in enumerate(row, start=1):
            if idx in exempt:
                continue
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                c[str(v)] += 1
    return c

def snapshot_values(wb):
    return {ws.title: _sheet_values(ws) for ws in wb.worksheets}

def verify_no_value_change(before, wb):
    problems = []
    for ws in wb.worksheets:
        base = before.get(ws.title)
        if base is None:
            problems.append(f"シート『{ws.title}』が元ファイルに無い(追加されている)")
            continue
        extra = _sheet_values(ws) - base
        if extra:
            problems.append(f"シート『{ws.title}』で値の変更/追加を検知: {list(extra.items())[:5]}")
    return problems

def process_checked(wb, log):
    before = snapshot_values(wb)
    process_wb(wb, log)
    problems = verify_no_value_change(before, wb)
    if problems:
        raise ValueError("安全チェック失敗（セル値の変更を検知したため中止）: " + " / ".join(problems))
    log("  ✔ 安全チェック合格：セルの値は不変（削除・書式のみ）")
    return wb

def process(path, log):
    return process_checked(openpyxl.load_workbook(path), log)

def resolve_keireki():
    if "--delete-keireki" in sys.argv: return True
    if "--keep-keireki" in sys.argv:   return False
    try:
        ans = input("「経歴」シートを削除しますか? [y=削除 / N=保持] : ").strip().lower()
    except EOFError:
        ans = ""
    return ans in ("y", "yes")

def processed_name(name):
    base, ext = os.path.splitext(name)
    return base + "_processed" + (ext or ".xlsx")

def run(inputs, delete_keireki, progress=None):
    global DELETE_KEIREKI
    DELETE_KEIREKI = delete_keireki
    def emit(s):
        (progress or (lambda x: print(x, flush=True)))(s)
    files = []; base = None
    for p in inputs:
        if os.path.isdir(p):
            base = p
            files += sorted(glob.glob(os.path.join(p, "*.xlsx")))
        elif p.lower().endswith(".xlsx"):
            files.append(p)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        emit("対象の .xlsx が見つかりませんでした。")
        return None, []
    if base is None:
        base = os.path.dirname(os.path.abspath(files[0]))
    out_dir = os.path.join(base, "_processed")
    logs_dir = os.path.join(base, "_logs")
    os.makedirs(out_dir, exist_ok=True); os.makedirs(logs_dir, exist_ok=True)
    emit("経歴シート: " + ("削除" if delete_keireki else "保持"))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    results = []
    for f in files:
        name = os.path.basename(f)
        if os.path.dirname(os.path.abspath(f)) == os.path.abspath(out_dir):
            continue
        lines = []
        def log(s, lines=lines): lines.append(s); emit(s)
        emit(f"\n===== {name} =====")
        try:
            wb = process(f, log)
            out_name = processed_name(name)
            wb.save(os.path.join(out_dir, out_name))
            log(f"保存 -> _processed/{out_name}")
            results.append((name, "OK"))
        except Exception as e:
            log(f"!! エラー: {e}")
            results.append((name, f"ERROR: {e}"))
        with open(os.path.join(logs_dir, f"{name}.{stamp}.log"), "w", encoding="utf-8") as fp:
            fp.write("\n".join(lines))
    ok = sum(1 for _, s in results if s == "OK")
    emit(f"\n完了: {ok}/{len(results)} ファイル処理。出力先: {out_dir}")
    return out_dir, results

def main():
    inputs = [a for a in sys.argv[1:] if not a.startswith("-")]
    if not inputs:
        inputs = [BASE]
    run(inputs, resolve_keireki())

if __name__ == "__main__":
    main()
